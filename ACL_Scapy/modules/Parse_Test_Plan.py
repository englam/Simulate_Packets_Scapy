from openpyxl import load_workbook


def get_excel_cell(test_plan_file,test_plan_sheet,search_keyword):
    def get_excel_cells(test_plan_file,test_plan_sheet,search_keyword):
        wb = load_workbook(test_plan_file)
        sheet = wb.get_sheet_by_name(test_plan_sheet)
        for row in sheet.rows:
            for col in row:
                if col.value == search_keyword:
                    return([col.column,col.row])
                
    if get_excel_cells(test_plan_file,test_plan_sheet,search_keyword) == None:
        print ("[*] search_keyword not found: %s" %(search_keyword))
        return (["0",0])
    return (get_excel_cells(test_plan_file,test_plan_sheet,search_keyword))


def write_excel_cell(test_plan_file,test_plan_sheet,fw_cell,fw_version,result_cell,test_result):
    wb = load_workbook(test_plan_file)
    sheet = wb.get_sheet_by_name(test_plan_sheet)
    sheet[fw_cell] = fw_version
    sheet[result_cell] = test_result
    wb.save(test_plan_file)

if __name__ == '__main__':
    _col,_row=get_excel_cell(test_plan_file,test_plan_sheet,search_keyword)
    pass
